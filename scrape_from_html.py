#!/usr/bin/env python3
"""
Scrape from manually saved HTML files
WORKFLOW:
1. Open browser: https://www.amazon.in/s?k=godrej&i=kitchen&rh=n%3A976442031%2Cp_123%3A5365053
2. Save page as HTML (Cmd+S)
3. Run: python3 scrape_from_html.py saved_page.html
"""

import sys
import re
import time
import random
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import requests


def parse_products_from_html(html_file):
    """Extract products from saved HTML file"""
    print(f"\n📄 Reading {html_file}...")
    
    with open(html_file, 'r', encoding='utf-8') as f:
        html = f.read()
    
    soup = BeautifulSoup(html, 'html.parser')
    
    products = []
    product_divs = soup.find_all('div', {'data-component-type': 's-search-result'})
    
    if not product_divs:
        product_divs = soup.find_all('div', {'data-asin': True, 'data-index': True})
    
    if not product_divs:
        product_divs = soup.find_all('div', {'data-asin': True})
        product_divs = [div for div in product_divs if div.get('data-asin', '').strip()]
    
    for product_div in product_divs:
        try:
            product_data = {}
            
            asin = product_div.get('data-asin', '')
            if not asin:
                continue
            product_data['asin'] = asin
            
            # Title
            title_elem = product_div.find('h2')
            if title_elem:
                title_link = title_elem.find('a')
                product_data['title'] = title_link.get_text(strip=True) if title_link else title_elem.get_text(strip=True)
            else:
                title_span = product_div.find('span', {'class': 'a-size-medium'})
                product_data['title'] = title_span.get_text(strip=True) if title_span else 'N/A'
            
            product_data['url'] = f'https://www.amazon.in/dp/{asin}'
            
            # Price
            price_whole = product_div.find('span', {'class': 'a-price-whole'})
            if price_whole:
                price_fraction = product_div.find('span', {'class': 'a-price-fraction'})
                price = price_whole.get_text(strip=True).replace(',', '')
                if price_fraction:
                    price += price_fraction.get_text(strip=True)
                product_data['price'] = f"₹{price}"
            else:
                product_data['price'] = 'N/A'
            
            # MRP
            mrp_elem = product_div.find('span', {'class': 'a-price a-text-price'})
            product_data['mrp'] = mrp_elem.get_text(strip=True) if mrp_elem else 'N/A'
            
            # Discount
            discount_elem = product_div.find('span', string=re.compile(r'\d+%\s+off'))
            product_data['discount'] = discount_elem.get_text(strip=True) if discount_elem else 'N/A'
            
            # Rating
            rating_elem = product_div.find('span', {'class': 'a-icon-alt'})
            if rating_elem:
                rating_text = rating_elem.get_text(strip=True)
                rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                product_data['rating'] = rating_match.group(1) if rating_match else 'N/A'
            else:
                product_data['rating'] = 'N/A'
            
            # Reviews
            reviews_elem = product_div.find('span', {'class': 'a-size-base'})
            if reviews_elem:
                reviews_text = reviews_elem.get_text(strip=True)
                reviews_match = re.search(r'\(([0-9.,KMk]+)\)', reviews_text)
                product_data['num_reviews'] = reviews_match.group(1) if reviews_match else 'N/A'
            else:
                product_data['num_reviews'] = 'N/A'
            
            products.append(product_data)
            
        except Exception as e:
            continue
    
    print(f"✓ Found {len(products)} products")
    return products


def get_brand(text):
    """Determine protection plan brand"""
    t = text.lower()
    if 'acko' in t: return 'Acko'
    if 'onsitego' in t: return 'Onsitego'
    if 'one assist' in t: return 'One Assist'
    if 'zopper' in t: return 'Zopper India'
    if 'servify' in t: return 'Servify'
    if ('warranty' in t or 'protection' in t) and 'by' not in t: 
        return 'Zopper India'
    return 'Unknown'


PLAN_PATTERN = re.compile(r'(.+?)\s+(?:for|at|from)\s+₹\s*([0-9,]+(?:\.\d{2})?)', re.IGNORECASE)


def extract_protection_plans(asin):
    """Extract protection plans - returns empty if can't access"""
    url = f'https://www.amazon.in/dp/{asin}'
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        }
        resp = requests.get(url, headers=headers, timeout=15)
        
        if len(resp.text) < 10000:
            return []  # Blocked
        
        soup = BeautifulSoup(resp.content, 'html.parser')
        plans = []
        seen = set()
        
        # Find protection section
        protection_section = None
        headings = soup.find_all(['h2', 'h3', 'h4', 'span'], string=re.compile(r'protection', re.IGNORECASE))
        if headings:
            protection_section = headings[0].find_parent(['div', 'section'])
        
        containers = (protection_section.find_all(['label', 'span', 'div']) if protection_section else [])
        for node in containers:
            text = node.get_text(' ', strip=True)
            m = PLAN_PATTERN.search(text)
            if not m:
                continue
            
            plan_name = m.group(1).strip()
            price_text = m.group(2)
            
            keywords = ['warranty','protection','plan','service','year','extended']
            if not any(k in plan_name.lower() for k in keywords):
                continue
            
            key = plan_name.lower()
            if key in seen or len(plan_name) < 8:
                continue
            seen.add(key)
            
            try:
                price = float(price_text.replace(',', ''))
            except:
                price = None
            
            brand = get_brand(plan_name)
            plans.append({
                'Plan Name': plan_name,
                'Plan Price': price,
                'Plan Brand': brand
            })
            
            if len(plans) >= 3:
                break
        
        return plans
        
    except Exception as e:
        return []


def extract_plans_for_all(products):
    """Extract plans - skips if blocked"""
    print("\n" + "="*80)
    print("STEP 2: EXTRACTING PROTECTION PLANS")
    print("="*80)
    print("\n⚠️  If product pages are blocked, plans will be empty")
    print("   (Product scraping still works from saved HTML)\n")
    
    results = []
    
    for idx, product in enumerate(products, 1):
        asin = product['asin']
        name = product['title']
        
        print(f"[{idx}/{len(products)}] {asin}... ", end="", flush=True)
        
        plans = extract_protection_plans(asin)
        
        # Ensure exactly 3 plan slots
        while len(plans) < 3:
            plans.append({'Plan Name': None, 'Plan Price': None, 'Plan Brand': None})
        plans = plans[:3]
        
        result = {
            'ASIN': asin,
            'Product Name': name,
            'Current Price': product.get('price'),
            'MRP': product.get('mrp'),
            'Discount %': product.get('discount'),
            'Rating': product.get('rating'),
            'Number of Reviews': product.get('num_reviews'),
            'Product Link': product.get('url'),
            'Plan 1 Name': plans[0]['Plan Name'],
            'Plan 1 Price': plans[0]['Plan Price'],
            'Plan 1 Brand': plans[0]['Plan Brand'],
            'Plan 2 Name': plans[1]['Plan Name'],
            'Plan 2 Price': plans[1]['Plan Price'],
            'Plan 2 Brand': plans[1]['Plan Brand'],
            'Plan 3 Name': plans[2]['Plan Name'],
            'Plan 3 Price': plans[2]['Plan Price'],
            'Plan 3 Brand': plans[2]['Plan Brand'],
        }
        
        results.append(result)
        
        plan_count = len([p for p in plans if p['Plan Name']])
        if plan_count > 0:
            print(f"✓ {plan_count} plan(s)")
        else:
            print("⚠ No plans")
        
        time.sleep(1.5 + random.random())
    
    return results


def create_excel(data, filename):
    """Create Excel with hyperlinks"""
    print("\n" + "="*80)
    print("STEP 3: CREATING EXCEL FILE")
    print("="*80 + "\n")
    
    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Godrej Products"
    
    headers = list(df.columns)
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    link_col_idx = headers.index('Product Link') + 1 if 'Product Link' in headers else None
    
    for idx, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        
        if link_col_idx and pd.notna(row['Product Link']):
            cell = ws.cell(row=idx+2, column=link_col_idx)
            cell.hyperlink = str(row['Product Link'])
            cell.font = Font(color="0563C1", underline="single")
            cell.value = "🔗 View on Amazon"
    
    widths = {'A': 15, 'B': 60, 'C': 15, 'D': 15, 'E': 12, 'F': 10, 'G': 18, 'H': 20,
              'I': 35, 'J': 12, 'K': 15, 'L': 35, 'M': 12, 'N': 15, 'O': 35, 'P': 12, 'Q': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(filename)
    print(f"✓ Saved: {filename}")
    
    try:
        import shutil
        desktop = '/Users/harshdeepsingh/Desktop/' + filename
        shutil.copy(filename, desktop)
        print(f"✓ Copied to Desktop!")
    except:
        pass


def main():
    print("\n" + "="*80)
    print("🚀 SCRAPE FROM SAVED HTML")
    print("="*80)
    print("\nWORKFLOW:")
    print("1. Open: https://www.amazon.in/s?k=godrej&i=kitchen&rh=n%3A976442031%2Cp_123%3A5365053")
    print("2. Save page (Cmd+S) as 'Webpage, Complete'")
    print("3. Run: python3 scrape_from_html.py <saved_file.html>")
    print("="*80)
    
    if len(sys.argv) < 2:
        print("\n❌ Please provide HTML file path")
        print(f"\nUsage: python3 {sys.argv[0]} <html_file>\n")
        return
    
    html_file = sys.argv[1]
    
    # Parse products from HTML
    products = parse_products_from_html(html_file)
    
    if not products:
        print("\n❌ No products found in HTML file")
        return
    
    # Extract protection plans
    results = extract_plans_for_all(products)
    
    # Create Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'godrej_from_html_{timestamp}.xlsx'
    create_excel(results, filename)
    
    # Summary
    print("\n" + "="*80)
    print("✅ DONE!")
    print("="*80)
    
    with_plans = sum(1 for r in results if r['Plan 1 Name'] is not None)
    total_plans = sum(1 for r in results for i in [1,2,3] if r.get(f'Plan {i} Name'))
    
    print(f"\n📊 SUMMARY:")
    print(f"   • Products scraped: {len(products)}")
    print(f"   • Products with plans: {with_plans}")
    print(f"   • Total protection plans: {total_plans}")
    print(f"\n📁 Output file: {filename}")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
