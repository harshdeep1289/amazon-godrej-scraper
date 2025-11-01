#!/usr/bin/env python3
"""
Add clickable hyperlinks to Product Link column in Excel file
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

print("\n" + "="*80)
print("ADDING HYPERLINKS TO EXCEL FILE")
print("="*80 + "\n")

# Load the existing Excel file
input_file = 'godrej_with_plans_20251029_033212.xlsx'
print(f"Loading: {input_file}")

df = pd.read_excel(input_file)
print(f"✓ Loaded {len(df)} products\n")

# Create new workbook with hyperlinks
print("Creating Excel file with clickable links...")

wb = Workbook()
ws = wb.active
ws.title = "Godrej Products with Plans"

# Write headers
headers = list(df.columns)
ws.append(headers)

# Style headers
for cell in ws[1]:
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Find Product Link column index
link_col_idx = headers.index('Product Link') + 1 if 'Product Link' in headers else None

# Write data rows with hyperlinks
for idx, row in df.iterrows():
    row_data = list(row)
    ws.append(row_data)
    
    # Add clickable hyperlink to Product Link cell
    if link_col_idx and pd.notna(row['Product Link']):
        cell = ws.cell(row=idx+2, column=link_col_idx)
        cell.hyperlink = str(row['Product Link'])
        cell.font = Font(color="0563C1", underline="single")
        cell.value = "🔗 View on Amazon"

# Adjust column widths for better readability
column_widths = {
    'A': 15,   # ASIN
    'B': 60,   # Product Name
    'C': 15,   # Current Price
    'D': 15,   # MRP
    'E': 12,   # Discount %
    'F': 10,   # Rating
    'G': 18,   # Number of Reviews
    'H': 20,   # Product Link
    'I': 35,   # Plan 1 Name
    'J': 12,   # Plan 1 Price
    'K': 15,   # Plan 1 Brand
    'L': 35,   # Plan 2 Name
    'M': 12,   # Plan 2 Price
    'N': 15,   # Plan 2 Brand
    'O': 35,   # Plan 3 Name
    'P': 12,   # Plan 3 Price
    'Q': 15,   # Plan 3 Brand
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save with new filename
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'godrej_with_plans_HYPERLINKED_{timestamp}.xlsx'

wb.save(output_file)

print(f"✓ Created: {output_file}")

# Copy to Desktop
import shutil
desktop_path = f'/Users/harshdeepsingh/Desktop/{output_file}'
shutil.copy(output_file, desktop_path)

print(f"✓ Copied to Desktop!")
print("\n" + "="*80)
print("✅ DONE! Open the file to see clickable Amazon links.")
print("="*80)
