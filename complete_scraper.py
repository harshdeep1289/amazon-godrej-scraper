#!/usr/bin/env python3
"""
Complete Amazon Godrej Product Scraper
Scrapes products, extracts protection plans, and creates a hyperlinked Excel file
Run this single script to get everything done!
"""

import random
import requests
from bs4 import BeautifulSoup
import re
import time
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from email_report import send_report

# Multiple user agents to rotate
USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
]

BASE_HEADERS = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,en-GB;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'DNT': '1',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Referer': 'https://www.amazon.in/',
}

# Cookie string - paste your browser cookies here after visiting Amazon
# To get cookies: Open browser DevTools (F12) > Application/Storage > Cookies > amazon.in
# Copy the cookie string and paste below
COOKIE_STRING = "session-id=523-6631916-7995726; ubid-acbin=523-5593219-3573351; sst-acbin=Sst1|PQGx9nWGogqZM8waKLvkRCzwB7NsKnR0Thqnh-i0WmKIg1mbfQf5Q2Vhx0mSkYiKc7RXZOytZcabBamnmhh8AwHbd5fpsG0dV-8nMVUGocQ5Ky1OLSuONOtieI1E8iD5_8bgAhCcns6Yq_kljC0MKxVURpBfqqZhcdFdcAc6aWv0_k05776wrnMOsUXzW0GjeLp_8K1Iyduzmth3n1mBFNHION8v73B1wjuPGjSwzAlpy0tyeiBDXe8haXfpbNr7Z40Q"

def make_headers():
    headers = {
        **BASE_HEADERS,
        'User-Agent': random.choice(USER_AGENTS),
    }
    if COOKIE_STRING:
        headers['Cookie'] = COOKIE_STRING
    return headers


def get_brand(text):
    """Determine protection plan brand from text"""
    t = text.lower()
    if 'acko' in t: return 'Acko'
    if 'onsitego' in t: return 'Onsitego'
    if 'one assist' in t: return 'One Assist'
    if 'zopper' in t: return 'Zopper India'
    if 'servify' in t: return 'Servify'
    if ('warranty' in t or 'protection' in t) and 'by' not in t: 
        return 'Zopper India'
    return 'Unknown'


PLAN_PATTERN = re.compile(r'(.+?)\s+(?:for|at|from)\s+₹\s*([0-9,]+(?:\.\d{2})?)', re.IGNORECASE)

def extract_protection_plans(asin, name, session=None):
    """Extract up to 3 protection plans from product page"""
    url = f'https://www.amazon.in/dp/{asin}'
    
    try:
        if session is None:
            session = requests.Session()
        
        resp = session.get(url, headers=make_headers(), timeout=20)
        
        # Check if blocked
        if len(resp.text) < 10000:
            return []  # Likely blocked
        
        resp.raise_for_status()
        
        soup = BeautifulSoup(resp.content, 'html.parser')
        
        plans = []
        seen = set()
        
        # Strategy 1: Find protection section
        protection_section = None
        headings = soup.find_all(['h2', 'h3', 'h4', 'span'], string=re.compile(r'Add a Protection Plan', re.IGNORECASE))
        if headings:
            protection_section = headings[0].find_parent(['div', 'section'])
        if not protection_section:
            protection_section = soup.find('div', {'id': re.compile(r'warranty|protection', re.IGNORECASE)})
        
        # Strategy 2: Search within section
        containers = (protection_section.find_all(['label', 'span', 'div']) if protection_section else [])
        for node in containers:
            text = node.get_text(' ', strip=True)
            m = PLAN_PATTERN.search(text)
            if not m:
                continue
            
            plan_name = m.group(1).strip()
            price_text = m.group(2)
            
            keywords = ['warranty','protection','plan','service','cleaning','installation','year','extended','damage']
            if not any(k in plan_name.lower() for k in keywords):
                continue
            
            key = plan_name.lower()
            if key in seen or len(plan_name) < 8:
                continue
            seen.add(key)
            
            try:
                price = float(price_text.replace(',', ''))
            except:
                price = None
            
            brand = get_brand(plan_name)
            plans.append({
                'Plan Name': plan_name,
                'Plan Price': price,
                'Plan Brand': brand
            })
            
            if len(plans) >= 3:
                break
        
        # Strategy 3: Broader search if still missing
        if len(plans) < 3:
            for m in PLAN_PATTERN.finditer(soup.get_text("\n")):
                plan_name = m.group(1).strip()
                price_text = m.group(2)
                
                if len(plan_name) < 8 or plan_name.lower() in seen:
                    continue
                
                keywords = ['warranty','protection','plan','service','cleaning','installation','year','extended','damage']
                if not any(k in plan_name.lower() for k in keywords):
                    continue
                
                seen.add(plan_name.lower())
                
                try:
                    price = float(price_text.replace(',', ''))
                except:
                    price = None
                
                brand = get_brand(plan_name)
                plans.append({
                    'Plan Name': plan_name,
                    'Plan Price': price,
                    'Plan Brand': brand
                })
                
                if len(plans) >= 3:
                    break
        
        return plans
        
    except Exception as e:
        return []


def get_next_page_url(soup, current_url):
    """Extract next page URL from pagination"""
    from urllib.parse import urljoin
    
    try:
        # Look for the "Next" button/link
        next_button = soup.find('a', {'class': 's-pagination-next'})
        if next_button and next_button.get('href'):
            return urljoin(current_url, next_button['href'])
        
        # Alternative: look for pagination link with specific aria-label
        next_link = soup.find('a', {'aria-label': 'Go to next page'})
        if next_link and next_link.get('href'):
            return urljoin(current_url, next_link['href'])
    except Exception as e:
        print(f"Error finding next page: {e}")
    
    return None


def scrape_products_from_page(url, session=None):
    """Scrape products from a single page"""
    try:
        if session is None:
            session = requests.Session()
        
        # Set cookies individually for better handling
        if COOKIE_STRING:
            for cookie in COOKIE_STRING.split('; '):
                if '=' in cookie:
                    key, value = cookie.split('=', 1)
                    session.cookies.set(key, value, domain='.amazon.in')
        
        response = session.get(url, headers=make_headers(), timeout=30)
        
        # Debug: check response
        print(f"[Status: {response.status_code}, Size: {len(response.text)} bytes] ", end="", flush=True)
        
        if response.status_code not in [200, 202]:
            print(f"⚠ Unexpected status code: {response.status_code}")
            return [], None
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Check for captcha/bot detection
        if 'validatecaptcha' in response.text.lower() or 'enter the characters' in response.text.lower() or 'robot check' in response.text.lower():
            print("⚠ Bot detection/Captcha detected. Try: 1) Using browser to access URL first, 2) Using VPN, 3) Waiting longer between requests")
            return [], None
        
        # Check if page is too small (likely blocked/error)
        if len(response.text) < 10000:
            print("⚠ Response too small, likely blocked or error page")
            return [], None
        
        # Find all product containers - try multiple strategies
        products = []
        product_divs = soup.find_all('div', {'data-component-type': 's-search-result'})
        
        # Fallback: try alternative selectors
        if not product_divs:
            product_divs = soup.find_all('div', {'data-asin': True, 'data-index': True})
        
        if not product_divs:
            product_divs = soup.find_all('div', {'data-asin': True})
            # Filter out non-product divs (empty ASINs)
            product_divs = [div for div in product_divs if div.get('data-asin', '').strip()]
        
        for product_div in product_divs:
            try:
                product_data = {}
                
                # Extract ASIN
                asin = product_div.get('data-asin', '')
                product_data['asin'] = asin
                
                # Extract product title
                title_elem = product_div.find('h2')
                if title_elem:
                    title_link = title_elem.find('a')
                    if title_link:
                        product_data['title'] = title_link.get_text(strip=True)
                    else:
                        product_data['title'] = title_elem.get_text(strip=True)
                else:
                    title_span = product_div.find('span', {'class': 'a-size-medium'}) or \
                                product_div.find('span', {'class': 'a-size-base-plus'})
                    product_data['title'] = title_span.get_text(strip=True) if title_span else 'N/A'
                
                # Create product URL using ASIN
                if asin:
                    product_data['url'] = f'https://www.amazon.in/dp/{asin}'
                else:
                    product_data['url'] = 'N/A'
                
                # Extract price
                price_whole = product_div.find('span', {'class': 'a-price-whole'})
                if price_whole:
                    price_fraction = product_div.find('span', {'class': 'a-price-fraction'})
                    price = price_whole.get_text(strip=True).replace(',', '')
                    if price_fraction:
                        price += price_fraction.get_text(strip=True)
                    product_data['price'] = f"₹{price}"
                else:
                    product_data['price'] = 'N/A'
                
                # Extract MRP (original price)
                mrp_elem = product_div.find('span', {'class': 'a-price a-text-price'})
                if mrp_elem:
                    product_data['mrp'] = mrp_elem.get_text(strip=True)
                else:
                    product_data['mrp'] = 'N/A'
                
                # Extract discount percentage
                discount_elem = product_div.find('span', string=re.compile(r'\d+%\s+off'))
                if discount_elem:
                    product_data['discount'] = discount_elem.get_text(strip=True)
                else:
                    product_data['discount'] = 'N/A'
                
                # Extract rating
                rating_elem = product_div.find('span', {'class': 'a-icon-alt'})
                if rating_elem:
                    rating_text = rating_elem.get_text(strip=True)
                    rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                    product_data['rating'] = rating_match.group(1) if rating_match else 'N/A'
                else:
                    product_data['rating'] = 'N/A'
                
                # Extract number of reviews
                reviews_elem = product_div.find('span', {'class': 'a-size-base'})
                if reviews_elem:
                    reviews_text = reviews_elem.get_text(strip=True)
                    reviews_match = re.search(r'\(([0-9.,KMk]+)\)', reviews_text)
                    product_data['num_reviews'] = reviews_match.group(1) if reviews_match else 'N/A'
                else:
                    product_data['num_reviews'] = 'N/A'
                
                products.append(product_data)
                
            except Exception as e:
                continue
        
        # Get next page URL
        next_page_url = get_next_page_url(soup, url)
        
        return products, next_page_url
        
    except requests.exceptions.RequestException as e:
        print(f"Error fetching page: {str(e)}")
        return [], None


def scrape_all_products(start_url, max_pages=10):
    """Scrape all pages with pagination"""
    print("\n" + "="*80)
    print("STEP 1: SCRAPING PRODUCTS")
    print("="*80 + "\n")
    
    all_products = []
    current_url = start_url
    page_num = 1
    session = requests.Session()  # Use session to maintain cookies
    
    while current_url and page_num <= max_pages:
        print(f"📄 Page {page_num}... ", end="", flush=True)
        
        products, next_url = scrape_products_from_page(current_url, session=session)
        
        if products:
            all_products.extend(products)
            print(f"✓ Found {len(products)} products (Total: {len(all_products)})")
        else:
            print("⚠ No products found")
        
        if not next_url:
            print("\n✓ Reached last page")
            break
        
        current_url = next_url
        page_num += 1
        
        # Be polite with delays (randomized)
        time.sleep(2.5 + random.random() * 2)
    
    return all_products


def extract_plans_for_all_products(products, session=None):
    """Extract protection plans for all scraped products"""
    print("\n" + "="*80)
    print("STEP 2: EXTRACTING PROTECTION PLANS")
    print("="*80)
    print(f"\nProcessing {len(products)} products...")
    print(f"Est. time: ~{len(products)*2.5/60:.1f} minutes\n")
    
    if session is None:
        session = requests.Session()
    
    results = []
    
    for idx, product in enumerate(products, 1):
        asin = product['asin']
        name = product['title']
        
        print(f"[{idx}/{len(products)}] {asin[:10]}... ", end="", flush=True)
        
        # Extract protection plans
        plans = extract_protection_plans(asin, name, session=session)
        
        # Ensure exactly 3 plan slots
        while len(plans) < 3:
            plans.append({'Plan Name': None, 'Plan Price': None, 'Plan Brand': None})
        plans = plans[:3]
        
        result = {
            'ASIN': asin,
            'Product Name': name,
            'Current Price': product.get('price'),
            'MRP': product.get('mrp'),
            'Discount %': product.get('discount'),
            'Rating': product.get('rating'),
            'Number of Reviews': product.get('num_reviews'),
            'Product Link': product.get('url'),
            
            'Plan 1 Name': plans[0]['Plan Name'],
            'Plan 1 Price': plans[0]['Plan Price'],
            'Plan 1 Brand': plans[0]['Plan Brand'],
            
            'Plan 2 Name': plans[1]['Plan Name'],
            'Plan 2 Price': plans[1]['Plan Price'],
            'Plan 2 Brand': plans[1]['Plan Brand'],
            
            'Plan 3 Name': plans[2]['Plan Name'],
            'Plan 3 Price': plans[2]['Plan Price'],
            'Plan 3 Brand': plans[2]['Plan Brand'],
        }
        
        results.append(result)
        
        plan_count = len([p for p in plans if p['Plan Name']])
        if plan_count > 0:
            print(f"✓ {plan_count} plan(s)")
        else:
            print("⚠ No plans")
        
        # Delay between requests (randomized)
        if idx < len(products):
            time.sleep(2 + random.random() * 1.5)
    
    return results


def create_hyperlinked_excel(data, filename):
    """Create Excel file with clickable hyperlinks"""
    print("\n" + "="*80)
    print("STEP 3: CREATING HYPERLINKED EXCEL FILE")
    print("="*80 + "\n")
    
    df = pd.DataFrame(data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Godrej Products"
    
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Find Product Link column index
    link_col_idx = headers.index('Product Link') + 1 if 'Product Link' in headers else None
    
    # Write data rows with hyperlinks
    for idx, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        
        # Add clickable hyperlink to Product Link cell
        if link_col_idx and pd.notna(row['Product Link']) and row['Product Link'] != 'N/A':
            cell = ws.cell(row=idx+2, column=link_col_idx)
            cell.hyperlink = str(row['Product Link'])
            cell.font = Font(color="0563C1", underline="single")
            cell.value = "🔗 View on Amazon"
    
    # Adjust column widths for better readability
    column_widths = {
        'A': 15,   # ASIN
        'B': 60,   # Product Name
        'C': 15,   # Current Price
        'D': 15,   # MRP
        'E': 12,   # Discount %
        'F': 10,   # Rating
        'G': 18,   # Number of Reviews
        'H': 20,   # Product Link
        'I': 35,   # Plan 1 Name
        'J': 12,   # Plan 1 Price
        'K': 15,   # Plan 1 Brand
        'L': 35,   # Plan 2 Name
        'M': 12,   # Plan 2 Price
        'N': 15,   # Plan 2 Brand
        'O': 35,   # Plan 3 Name
        'P': 12,   # Plan 3 Price
        'Q': 15,   # Plan 3 Brand
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save file
    wb.save(filename)
    print(f"✓ Saved: {filename}")
    
    # Copy to Desktop
    try:
        import shutil
        desktop_path = f'/Users/harshdeepsingh/Desktop/{filename}'
        shutil.copy(filename, desktop_path)
        print(f"✓ Copied to Desktop!")
    except:
        pass


def main():
    """Main execution function"""
    print("\n" + "="*80)
    print("🚀 COMPLETE AMAZON GODREJ SCRAPER")
    print("="*80)
    print("This script will:")
    print("  1. Scrape all Godrej products from Amazon")
    print("  2. Extract protection plans for each product")
    print("  3. Create a hyperlinked Excel file")
    print("="*80)
    
    # Check for cookies
    if not COOKIE_STRING:
        print("\n⚠️  WARNING: No cookies set!")
        print("\nAmazon is blocking automated requests. To fix:")
        print("1. Open the URL below in your browser:")
        print("   https://www.amazon.in/s?k=godrej&i=kitchen")
        print("2. Open DevTools (F12 or Cmd+Option+I)")
        print("3. Go to Application/Storage tab > Cookies > amazon.in")
        print("4. Copy ALL cookies as a string (or use EditThisCookie extension)")
        print("5. Paste the cookie string in complete_scraper.py at line ~40")
        print("   Example: COOKIE_STRING = \"session-id=123; ubid-acbin=456...\"")
        print("\nTrying without cookies (may fail)...\n")
        time.sleep(3)
    
    # Amazon search URL for Godrej kitchen appliances
    url = "https://www.amazon.in/s?k=godrej&i=kitchen&rh=n%3A976442031%2Cp_123%3A5365053&dc&ds=v1%3AJs8a30i4FfROEbRBB51hMH89R8xeXGCWa0bHe%2FcXs10&qid=1761798762&rnid=91049095031&ref=sr_nr_p_123_3"
    
    print(f"\n🔗 Search URL: {url}")
    print("\n⚠️  NOTE: If scraping fails, Amazon may be blocking automated requests.")
    print("   Try: 1) Visit the URL in your browser first to establish cookies")
    print("        2) Use a VPN or different network")
    print("        3) Increase delay between requests\n")
    
    # Create session for maintaining cookies
    session = requests.Session()
    
    # Step 1: Scrape products
    products = scrape_all_products(url, max_pages=20)
    
    if not products:
        print("\n❌ No products scraped. Exiting.")
        return
    
    # Step 2: Extract protection plans (use same session)
    results = extract_plans_for_all_products(products, session=session)
    
    # Step 3: Create Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'godrej_complete_{timestamp}.xlsx'
    create_hyperlinked_excel(results, filename)

    # Step 4: Email report (if configured via env)
    try:
        sent = send_report(
            filename,
            subject=f"Godrej Report {timestamp}",
            body="Automated report attached.",
        )
        if sent:
            print("✓ Report emailed")
    except Exception as e:
        print(f"⚠️ Email step skipped: {e}")
    
    # Final summary
    print("\n" + "="*80)
    print("✅ ALL DONE!")
    print("="*80)
    
    with_plans = sum(1 for r in results if r['Plan 1 Name'] is not None)
    total_plans = sum(1 for r in results for i in [1,2,3] if r.get(f'Plan {i} Name'))
    
    print(f"\n📊 SUMMARY:")
    print(f"   • Products scraped: {len(products)}")
    print(f"   • Products with plans: {with_plans}")
    print(f"   • Total protection plans: {total_plans}")
    print(f"\n📁 Output file: {filename}")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
