#!/usr/bin/env python3
"""
Complete Selenium-Based Amazon Scraper
Bypasses bot detection by using real browser automation
"""

import time
import random
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
except ImportError:
    print("❌ Selenium not installed!")
    print("\nInstall with:")
    print("  pip3 install selenium")
    print("\nThen install ChromeDriver:")
    print("  brew install chromedriver")
    exit(1)


def setup_driver():
    """Setup Chrome driver with stealth settings"""
    options = Options()
    
    # Stealth settings to avoid detection
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # User agent
    options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    # Optional: run headless
    # options.add_argument('--headless')
    
    try:
        driver = webdriver.Chrome(options=options)
        # Remove webdriver property
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        return driver
    except Exception as e:
        print(f"❌ Error setting up Chrome driver: {e}")
        print("\nMake sure ChromeDriver is installed:")
        print("  brew install chromedriver")
        exit(1)


def get_brand(text):
    """Determine protection plan brand"""
    t = text.lower()
    if 'acko' in t: return 'Acko'
    if 'onsitego' in t: return 'Onsitego'
    if 'one assist' in t: return 'One Assist'
    if 'zopper' in t: return 'Zopper India'
    if 'servify' in t: return 'Servify'
    if ('warranty' in t or 'protection' in t) and 'by' not in t: 
        return 'Zopper India'
    return 'Unknown'


PLAN_PATTERN = re.compile(r'(.+?)\s+(?:for|at|from)\s+₹\s*([0-9,]+(?:\.\d{2})?)', re.IGNORECASE)


def extract_plans(driver, asin):
    """Extract protection plans from product page"""
    url = f'https://www.amazon.in/dp/{asin}'
    plans = []
    
    try:
        driver.get(url)
        time.sleep(2 + random.random())
        
        # Get page source and parse
        html = driver.page_source
        seen = set()
        
        for m in PLAN_PATTERN.finditer(html):
            plan_name = m.group(1).strip()
            plan_name = re.sub(r'<[^>]+>', '', plan_name)
            plan_name = re.sub(r'\s+', ' ', plan_name)
            price_text = m.group(2)
            
            if len(plan_name) < 8 or plan_name.lower() in seen:
                continue
            
            keywords = ['warranty','protection','plan','service','cleaning','installation','year','extended','damage']
            if not any(k in plan_name.lower() for k in keywords):
                continue
            
            seen.add(plan_name.lower())
            
            try:
                price = float(price_text.replace(',', ''))
            except:
                price = None
            
            brand = get_brand(plan_name)
            plans.append({
                'Plan Name': plan_name,
                'Plan Price': price,
                'Plan Brand': brand
            })
            
            if len(plans) >= 3:
                break
        
        return plans
    except Exception as e:
        return []


def scrape_page(driver, url):
    """Scrape products from a page"""
    try:
        driver.get(url)
        time.sleep(3 + random.random() * 2)
        
        # Wait for products to load
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-component-type="s-search-result"]')))
        
        products = []
        product_elements = driver.find_elements(By.CSS_SELECTOR, '[data-component-type="s-search-result"]')
        
        for elem in product_elements:
            try:
                asin = elem.get_attribute('data-asin')
                if not asin:
                    continue
                
                # Title
                try:
                    title = elem.find_element(By.CSS_SELECTOR, 'h2 a span').text
                except:
                    title = 'N/A'
                
                # Price
                try:
                    price = elem.find_element(By.CSS_SELECTOR, '.a-price-whole').text
                    price = f"₹{price}"
                except:
                    price = 'N/A'
                
                # MRP
                try:
                    mrp = elem.find_element(By.CSS_SELECTOR, '.a-price.a-text-price').text
                except:
                    mrp = 'N/A'
                
                # Discount
                try:
                    discount = elem.find_element(By.XPATH, ".//*[contains(text(), '% off')]").text
                except:
                    discount = 'N/A'
                
                # Rating
                try:
                    rating = elem.find_element(By.CSS_SELECTOR, '.a-icon-alt').text
                    rating = re.search(r'(\d+\.?\d*)', rating).group(1) if rating else 'N/A'
                except:
                    rating = 'N/A'
                
                # Reviews
                try:
                    reviews = elem.find_element(By.CSS_SELECTOR, 'span.a-size-base').text
                    reviews_match = re.search(r'\(([0-9.,KMk]+)\)', reviews)
                    reviews = reviews_match.group(1) if reviews_match else 'N/A'
                except:
                    reviews = 'N/A'
                
                products.append({
                    'asin': asin,
                    'title': title,
                    'price': price,
                    'mrp': mrp,
                    'discount': discount,
                    'rating': rating,
                    'num_reviews': reviews,
                    'url': f'https://www.amazon.in/dp/{asin}'
                })
                
            except Exception as e:
                continue
        
        # Check for next page
        try:
            next_button = driver.find_element(By.CSS_SELECTOR, '.s-pagination-next')
            next_url = next_button.get_attribute('href')
        except:
            next_url = None
        
        return products, next_url
        
    except Exception as e:
        print(f"Error scraping page: {e}")
        return [], None


def scrape_all_products(driver, start_url, max_pages=20):
    """Scrape all pages"""
    print("\n" + "="*80)
    print("STEP 1: SCRAPING PRODUCTS")
    print("="*80 + "\n")
    
    all_products = []
    current_url = start_url
    page_num = 1
    
    while current_url and page_num <= max_pages:
        print(f"📄 Page {page_num}... ", end="", flush=True)
        
        products, next_url = scrape_page(driver, current_url)
        
        if products:
            all_products.extend(products)
            print(f"✓ Found {len(products)} products (Total: {len(all_products)})")
        else:
            print("⚠ No products")
            break
        
        if not next_url:
            print("\n✓ Reached last page")
            break
        
        current_url = next_url
        page_num += 1
        time.sleep(3 + random.random() * 2)
    
    return all_products


def extract_plans_for_all(driver, products):
    """Extract plans for all products"""
    print("\n" + "="*80)
    print("STEP 2: EXTRACTING PROTECTION PLANS")
    print("="*80)
    print(f"\nProcessing {len(products)} products...\n")
    
    results = []
    
    for idx, product in enumerate(products, 1):
        asin = product['asin']
        name = product['title']
        
        print(f"[{idx}/{len(products)}] {asin}... ", end="", flush=True)
        
        plans = extract_plans(driver, asin)
        
        # Ensure exactly 3 plan slots
        while len(plans) < 3:
            plans.append({'Plan Name': None, 'Plan Price': None, 'Plan Brand': None})
        plans = plans[:3]
        
        result = {
            'ASIN': asin,
            'Product Name': name,
            'Current Price': product.get('price'),
            'MRP': product.get('mrp'),
            'Discount %': product.get('discount'),
            'Rating': product.get('rating'),
            'Number of Reviews': product.get('num_reviews'),
            'Product Link': product.get('url'),
            'Plan 1 Name': plans[0]['Plan Name'],
            'Plan 1 Price': plans[0]['Plan Price'],
            'Plan 1 Brand': plans[0]['Plan Brand'],
            'Plan 2 Name': plans[1]['Plan Name'],
            'Plan 2 Price': plans[1]['Plan Price'],
            'Plan 2 Brand': plans[1]['Plan Brand'],
            'Plan 3 Name': plans[2]['Plan Name'],
            'Plan 3 Price': plans[2]['Plan Price'],
            'Plan 3 Brand': plans[2]['Plan Brand'],
        }
        
        results.append(result)
        
        plan_count = len([p for p in plans if p['Plan Name']])
        if plan_count > 0:
            print(f"✓ {plan_count} plan(s)")
        else:
            print("⚠ No plans")
        
        time.sleep(2 + random.random())
    
    return results


def create_excel(data, filename):
    """Create Excel with hyperlinks"""
    print("\n" + "="*80)
    print("STEP 3: CREATING EXCEL FILE")
    print("="*80 + "\n")
    
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Godrej Products"
    
    headers = list(df.columns)
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    link_col_idx = headers.index('Product Link') + 1 if 'Product Link' in headers else None
    
    for idx, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        
        if link_col_idx and pd.notna(row['Product Link']):
            cell = ws.cell(row=idx+2, column=link_col_idx)
            cell.hyperlink = str(row['Product Link'])
            cell.font = Font(color="0563C1", underline="single")
            cell.value = "🔗 View on Amazon"
    
    # Column widths
    widths = {'A': 15, 'B': 60, 'C': 15, 'D': 15, 'E': 12, 'F': 10, 'G': 18, 'H': 20,
              'I': 35, 'J': 12, 'K': 15, 'L': 35, 'M': 12, 'N': 15, 'O': 35, 'P': 12, 'Q': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(filename)
    print(f"✓ Saved: {filename}")
    
    try:
        import shutil
        desktop = f'/Users/harshdeepsingh/Desktop/{filename}'
        shutil.copy(filename, desktop)
        print(f"✓ Copied to Desktop!")
    except:
        pass


def main():
    print("\n" + "="*80)
    print("🚀 SELENIUM-BASED AMAZON SCRAPER")
    print("="*80)
    print("This uses real browser automation to bypass bot detection")
    print("="*80)
    
    url = "https://www.amazon.in/s?k=godrej&i=kitchen&rh=n%3A976442031%2Cp_123%3A5365053&dc&ds=v1%3AJs8a30i4FfROEbRBB51hMH89R8xeXGCWa0bHe%2FcXs10&qid=1761798762&rnid=91049095031&ref=sr_nr_p_123_3"
    
    # Setup driver
    print("\n🌐 Opening Chrome browser...")
    driver = setup_driver()
    
    try:
        # Scrape products
        products = scrape_all_products(driver, url, max_pages=20)
        
        if not products:
            print("\n❌ No products scraped. Exiting.")
            return
        
        # Extract plans
        results = extract_plans_for_all(driver, products)
        
        # Create Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'godrej_complete_{timestamp}.xlsx'
        create_excel(results, filename)
        
        # Summary
        print("\n" + "="*80)
        print("✅ ALL DONE!")
        print("="*80)
        
        with_plans = sum(1 for r in results if r['Plan 1 Name'] is not None)
        total_plans = sum(1 for r in results for i in [1,2,3] if r.get(f'Plan {i} Name'))
        
        print(f"\n📊 SUMMARY:")
        print(f"   • Products scraped: {len(products)}")
        print(f"   • Products with plans: {with_plans}")
        print(f"   • Total protection plans: {total_plans}")
        print(f"\n📁 Output file: {filename}")
        print("="*80 + "\n")
        
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
