#!/usr/bin/env python3
"""
Working Amazon Protection Plans Extractor
Properly extracts warranty plans with prices from Amazon product pages
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
from datetime import datetime

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
}

def get_brand(text):
    """Determine brand from plan text"""
    t = text.lower()
    if 'acko' in t: return 'Acko'
    if 'onsitego' in t: return 'Onsitego'
    if 'one assist' in t: return 'One Assist'
    if 'zopper' in t: return 'Zopper India'
    if 'servify' in t: return 'Servify'
    # Default if no "by" mentioned
    if ('warranty' in t or 'protection' in t) and 'by' not in t: 
        return 'Zopper India'
    return 'Unknown'

def extract_plans(asin, name):
    """Extract protection plans from product page"""
    url = f'https://www.amazon.in/dp/{asin}'
    
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        
        soup = BeautifulSoup(resp.content, 'html.parser')
        
        plans = []
        seen = set()
        
        # Method 1: Look for warranty checkboxes/labels
        # Amazon uses labels for warranty options
        labels = soup.find_all('label')
        
        for label in labels:
            text = label.get_text(" ", strip=True)
            
            # Look for warranty/plan keywords
            if not any(k in text.lower() for k in ['warranty', 'protection', 'plan', 'year']):
                continue
            
            # Find price within the label
            price_span = label.find('span', class_=re.compile('price'))
            if not price_span:
                continue
            
            price_text = price_span.get_text(strip=True)
            
            # Extract price value
            price_match = re.search(r'₹\s*([0-9,]+(?:\.\d{2})?)', price_text)
            if not price_match:
                continue
            
            price = float(price_match.group(1).replace(',', ''))
            
            # Clean plan name (remove price part)
            plan_name = re.sub(r'\s+for\s+₹.*$', '', text, flags=re.IGNORECASE).strip()
            plan_name = re.sub(r'₹.*$', '', plan_name).strip()
            
            # Skip duplicates
            key = plan_name.lower()
            if key in seen or len(plan_name) < 10:
                continue
            
            seen.add(key)
            
            brand = get_brand(plan_name)
            
            # Also check parent div for brand info
            parent_text = label.find_parent('div').get_text(" ", strip=True) if label.find_parent('div') else ""
            if 'from' in parent_text.lower():
                brand_match = re.search(r'from\s+([A-Za-z\s]+)', parent_text, re.IGNORECASE)
                if brand_match:
                    extracted_brand = brand_match.group(1).strip()
                    if extracted_brand:
                        brand = extracted_brand.title()
            
            plans.append({
                'Plan Name': plan_name,
                'Plan Price': price,
                'Plan Brand': brand
            })
            
            if len(plans) >= 3:
                break
        
        #Method 2: If no plans found, try broader search
        if not plans:
            # Search entire page for warranty text with prices
            all_text = soup.get_text()
            # Look for patterns like "2Year Extended Warranty Plan for ₹899"
            pattern = r'([\w\s]{15,80}(?:warranty|protection|plan)[\w\s]{0,30})[\s\n]*for[\s\n]*₹\s*([0-9,]+(?:\.\d{2})?)'
            matches = re.findall(pattern, all_text, re.IGNORECASE)
            
            for plan_text, price_text in matches:
                plan_name = ' '.join(plan_text.split())
                
                key = plan_name.lower()
                if key in seen or len(plan_name) < 10:
                    continue
                
                seen.add(key)
                
                price = float(price_text.replace(',', ''))
                brand = get_brand(plan_name)
                
                plans.append({
                    'Plan Name': plan_name,
                    'Plan Price': price,
                    'Plan Brand': brand
                })
                
                if len(plans) >= 3:
                    break
        
        return plans
        
    except Exception as e:
        return []

def main():
    print("\n" + "="*80)
    print("AMAZON PROTECTION PLANS EXTRACTOR")
    print("="*80)
    
    try:
        df = pd.read_excel('amazon_godrej_products.xlsx')
        print(f"\n✓ Loaded {len(df)} products")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        return
    
    print(f"\nExtracting protection plans from {len(df)} products...")
    print(f"Est. time: ~{len(df)*2.5/60:.1f} minutes (with 2s delays)")
    print("="*80 + "\n")
    
    results = []
    
    for idx, row in df.iterrows():
        asin = row['ASIN']
        name = row['Product Name']
        
        print(f"[{idx+1}/{len(df)}] {asin}...", end=" ", flush=True)
        
        plans = extract_plans(asin, name)
        
        # Ensure exactly 3 plan slots
        while len(plans) < 3:
            plans.append({'Plan Name': None, 'Plan Price': None, 'Plan Brand': None})
        plans = plans[:3]
        
        result = {
            'ASIN': asin,
            'Product Name': name,
            'Current Price': row.get('Current Price'),
            'MRP': row.get('MRP'),
            'Discount %': row.get('Discount %'),
            'Rating': row.get('Rating'),
            'Number of Reviews': row.get('Number of Reviews'),
            'Product Link': row.get('Product Link'),
            
            'Plan 1 Name': plans[0]['Plan Name'],
            'Plan 1 Price': plans[0]['Plan Price'],
            'Plan 1 Brand': plans[0]['Plan Brand'],
            
            'Plan 2 Name': plans[1]['Plan Name'],
            'Plan 2 Price': plans[1]['Plan Price'],
            'Plan 2 Brand': plans[1]['Plan Brand'],
            
            'Plan 3 Name': plans[2]['Plan Name'],
            'Plan 3 Price': plans[2]['Plan Price'],
            'Plan 3 Brand': plans[2]['Plan Brand'],
        }
        
        results.append(result)
        
        plan_count = len([p for p in plans if p['Plan Name']])
        if plan_count > 0:
            print(f"✓ {plan_count} plan(s)")
        else:
            print("⚠ No plans")
        
        # Delay between requests
        if idx + 1 < len(df):
            time.sleep(2)
    
    # Save results
    output_df = pd.DataFrame(results)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'godrej_with_plans_{timestamp}.xlsx'
    
    # Create Excel file with hyperlinks
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Godrej Products with Plans"
    
    # Write headers
    headers = list(output_df.columns)
    ws.append(headers)
    
    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Find Product Link column index
    link_col_idx = headers.index('Product Link') + 1 if 'Product Link' in headers else None
    
    # Write data rows
    for idx, row in output_df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        
        # Add hyperlink to Product Link cell
        if link_col_idx and row['Product Link']:
            cell = ws.cell(row=idx+2, column=link_col_idx)
            cell.hyperlink = row['Product Link']
            cell.font = Font(color="0563C1", underline="single")
            cell.value = "View Product"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # ASIN
    ws.column_dimensions['B'].width = 50  # Product Name
    ws.column_dimensions['H'].width = 20  # Product Link
    
    wb.save(output_file)
    
    print("\n" + "="*80)
    print("✅ EXTRACTION COMPLETE!")
    print("="*80)
    
    # Stats
    with_plans = sum(1 for r in results if r['Plan 1 Name'] is not None)
    total_plans = sum(1 for r in results for i in [1,2,3] if r.get(f'Plan {i} Name'))
    
    print(f"\nProducts processed: {len(results)}")
    print(f"Products with plans: {with_plans}")
    print(f"Total protection plans found: {total_plans}")
    
    print(f"\n✓ Saved to: {output_file}")
    
    # Copy to desktop
    import shutil
    desktop_path = f'/Users/harshdeepsingh/Desktop/{output_file}'
    shutil.copy(output_file, desktop_path)
    print(f"✓ Copied to Desktop!")
    print("="*80)

if __name__ == "__main__":
    main()
